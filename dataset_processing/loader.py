import argparse
import os
import json
from ctypes import Union
from pathlib import Path
import pandas as pd
import numpy as np
from datetime import datetime as dt
from dataset_processing.utils.annotate import ThermalIP
from openpyxl import Workbook as wb
from openpyxl import load_workbook
from datapoint import DataPoint
import cv2
from constants import Annotation 

DATASET_PATH = "test_data"
TEST_DATA_PATH = "../test_data_points"
PROCESSED_DATASET_PATH = "../test_data_processed"
DATASET_JSON = "dataset.json"
EXCEL_DATA_PATH = "thermal_collection.xlsx"
SHEET_NAME = "Annotations"

class DataLoader:
    def __init__(self, input_path=DATASET_PATH, output_path=PROCESSED_DATASET_PATH):
        # paths
        self.input_path = input_path
        self.output_path = output_path
        self.training_list = []
        self.testing_list = []

        # excel data
        self.df_collection = None
        self.df_windows = None

    def read_json(self):
        """
            takes the json file of dataset and reads the data
        :return:
            dictionary of the dataset. with list of data points
        """
        with open(DATASET_JSON, 'r') as f:
            data = json.load(f)
            self.training_list = data['train']
            self.testing_list = data['test']
            if data is None:
                self.load()

    def save_json(self):
        """ save the training data into json for faster retrieval"""
        with open(DATASET_JSON, "w") as outfile:
            json.dump({'train': self.training_list, 'test': self.testing_list}, outfile)

    def read_excel(self):
        """
            Reads the excel data collection and windows information into dataframes
        :return:
            pandas dataframe
        """
        self.df_collection = pd.read_excel(EXCEL_DATA_PATH, sheet_name='Collection')
        self.df_windows = pd.read_excel(EXCEL_DATA_PATH, sheet_name='Windows')

    def load(self):
        """
            takes the raw dataset directory as input and starts generating datapoints
            Returns:
                    None
        """
        if os.path.exists(DATASET_JSON):
            self.read_json()
            for index in range(len(self.training_list)):
                dp = self.training_list[index]
                self.training_list[index] = self.expand_data_point(dp)
            for index in range(len(self.testing_list)):
                dp = self.testing_list[index]
                self.testing_list[index] = self.expand_data_point(dp)
            return

        if not os.path.exists(EXCEL_DATA_PATH):
            print("Excel data sheet not found.")
            return
        else:
            self.read_excel()

        # walking through the data from E50
        for folder_name, _, filenames in os.walk(self.input_path):
            for filename in filenames:
                if filename.__contains__("IR"):
                    dp = {}
                    thermal_image_path = os.path.join(folder_name, filename)
                    dp["thermal_img_name"] = filename
                    dp['thermal_img_path'] = thermal_image_path
                    self.training_list.append(dp)
        self.save_json()
        self.load()

    def expand_data_point(self, dp):
        """takes json data point and generates numpy arrays into it
            :return
                dp with more attributes
        """
        data_point = DataPoint(dp['thermal_img_path'])
        dp['raw_thermal'] = data_point.get_thermal  # grayscale palette thermal image
        dp['raw_visual'] = data_point.get_visual
        dp['thermal_grayscale'] = cv2.cvtColor(dp['raw_thermal'], cv2.COLOR_BGR2GRAY)
        dp['fahrenheit'] = data_point.thermogram.fahrenheit
        # todo to call the aligner
        dp['aligned_visual'] = ""
        dp['date_taken'] = data_point.get_date
        # todo get the indoor and outdoor temp from the methods
        dp['indoor_temperature'] = 68
        dp['outdoor_temperature'] = 64
        dp['leak_info'] = data_point.get_opening
        dp['mask'] = None
        dp['window_name'] = data_point.get_window

        return dp

    def get_indoor_temperature(self):
        """
            gets the indoor temperature based on house and date
        :return:
            indoor temperature value
        """
        return self.df_collection.loc[
            (self.df_collection['House'] == 1 and dt(self.df_collection['Date']).date() == dt(self.cm.get_date).date()),'TempFInsideStart']

    def get_outdoor_temperature(self):
        """
            gets the outdoor temperature based on house and date
        :return:
            outdoor temperture value
        """
        #todo get the outdoor temperature values from Weather.com
        return self.df_collection.loc[
            (self.df_collection['House'] == 1 & self.df_collection['Date'] == self.cm.get_date.date()), 'TempFOutsideStart']

    def construct_annotation(self, dp) -> list:
        data_point = DataPoint(dp["thermal_img_path"])
        annotation = {}
        try:
            annotation[Annotation.IMAGE_NAME.value] = dp["thermal_img_name"]
            annotation[Annotation.DATE_TAKEN.value] = data_point.get_date
            annotation[Annotation.LEAK_SIGNATURE.value] = data_point.get_opening
            annotation[Annotation.WINDOW_NAME.value] = data_point.get_window
            annotation[Annotation.INSIDE_TEMP.value] = 68
            annotation[Annotation.OUTSIDE_TEMP.value] = 55
            annotation[Annotation.MIN_TEMP.value] = int(np.min(dp["fahrenheit"]))
            annotation[Annotation.MAX_TEMP.value] = int(np.max(dp["fahrenheit"]))
            print(annotation)
            annotation = dict(sorted(annotation.items()))
        except:
            print("error while contructing annotation")
        return [*annotation.values()]

    def update_annotation_sheet(self, xlsx_path = EXCEL_DATA_PATH) -> bool:

        workbook = load_workbook(filename=xlsx_path)
        try:
            # access the sheet if already exists
            if SHEET_NAME in workbook.sheetnames:
                sheet = workbook[SHEET_NAME]
                print(f"Accessing sheet: {sheet}")
            else:
                sheet = workbook.create_sheet(title  = SHEET_NAME)
            
            header_row = 1 # first row is the header row
            for index in range(len(self.training_list)):
                dp = self.training_list[index]
                found = False
                for row in sheet.iter_rows(values_only=True):
                    if dp["thermal_img_name"] in row:
                        found = True
                        break
                if not found:
                    annotation = self.construct_annotation(dp)
                    sheet.append(annotation)
        except Exception as e:
            print("error")
            print("Saving incomplete annotation sheet")
        finally:
            workbook.save(filename=xlsx_path)
            print("Saved Succesfully")


    def annotate(self):
        """
            performs bunchof image processing techniques so that the resulting image is ready for
            annotation
        Returns:
            data point with image processed information
        """
        # thermal_ip = ThermalIP(self.training_list)
        self.update_annotation_sheet()
