"""
Script to generate excel sheet with required data
"""
from openpyxl import Workbook as wb

DEFAULT_PATH = "dataset.xlsx"
SHEET_NAME = "Annotations"

def create_sheet(xlsx_path = DEFAULT_PATH) -> bool:
    workbook = wb.load_workbook(filename=xlsx_path)

    # access the sheet if already exists
    if SHEET_NAME in workbook.sheetnames:
        sheet = workbook[SHEET_NAME]
        print(f"Accessing sheet: {sheet}")
    else:
        sheet = workbook.create_sheet(title  = SHEET_NAME)
    
    header_row = 1 # first row is the header row
    



